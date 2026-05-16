"""xlsx → Supabase monitoring_sites 직접 upsert (REST API).

배경: AUSUHAK_monitoring_sites_v2.xlsx 12 시트 · 345 사이트 (📌 목차 시트 제외).
각 데이터 시트 구조:
  row 1   = 제목 (skip)
  row 2   = empty (skip)
  row 3   = 헤더 "카테고리/사이트명/URL/설명" (skip)
  row 4+  = 데이터 또는 섹션 헤더
    섹션 헤더 = 첫 컬럼이 "  ▣..." 이고 URL 비어있음 → section 으로 추적, 다음 데이터들에 상속
    데이터    = URL 존재 + 사이트명 존재

서비스 롤 키 사용 (RLS 우회). 배치 50개씩 → 진행률 출력.
ON CONFLICT (sheet, name, url) merge.

사용:
  python scripts/apply_monitoring_sites.py
환경:
  .env.local 에 NEXT_PUBLIC_SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY 필요.
"""
from __future__ import annotations

import re
import sys
import json
import urllib.request
import urllib.error
from pathlib import Path

import openpyxl

XLSX_PATH = Path("C:/Users/Wilson/Desktop/ausuhak_v25/04_ASSETS/AUSUHAK_monitoring_sites_v2.xlsx")
ENV_PATH = Path(".env.local")

# 📌 목차는 데이터 X — 스킵. 데이터 시트는 항상 이름이 "숫자." 로 시작.
DATA_SHEET_PREFIX_RE = re.compile(r"^\d+\.")


def clean(s: object) -> str:
    if s is None:
        return ""
    return str(s).strip()


def make_search_text(name: str, description: str, category: str, section: str) -> str:
    parts = [name, description, category, section]
    s = " ".join(p for p in parts if p)
    s = re.sub(r"\s+", " ", s).strip()
    return s[:2000]


def load_env() -> tuple[str, str]:
    if not ENV_PATH.exists():
        print(f"ERROR: {ENV_PATH} 없음", file=sys.stderr)
        sys.exit(1)
    env = {}
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    url = env.get("NEXT_PUBLIC_SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: NEXT_PUBLIC_SUPABASE_URL 또는 SUPABASE_SERVICE_ROLE_KEY 누락", file=sys.stderr)
        sys.exit(1)
    return url, key


def post(url: str, key: str, path: str, body: list | dict | None, prefer: str = "") -> tuple[int, dict | str, dict]:
    req = urllib.request.Request(
        f"{url}{path}",
        data=json.dumps(body).encode("utf-8") if body is not None else None,
        method="POST" if body is not None else "GET",
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": prefer,
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            raw = resp.read().decode("utf-8")
            try:
                data = json.loads(raw) if raw else {}
            except json.JSONDecodeError:
                data = raw
            return resp.status, data, dict(resp.headers)
    except urllib.error.HTTPError as e:
        raw = e.read().decode("utf-8", errors="replace")
        return e.code, raw, dict(e.headers)


def parse_xlsx() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    rows: list[dict] = []

    for sheet_name in wb.sheetnames:
        if not DATA_SHEET_PREFIX_RE.match(sheet_name):
            continue
        ws = wb[sheet_name]
        current_section = ""
        order = 0

        for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if r_idx <= 3:
                # 1=제목, 2=공백, 3=헤더
                continue
            if not row or all(c is None for c in row):
                continue

            col_cat = clean(row[0] if len(row) > 0 else None)
            col_name = clean(row[1] if len(row) > 1 else None)
            col_url = clean(row[2] if len(row) > 2 else None)
            col_desc = clean(row[3] if len(row) > 3 else None)

            # 섹션 헤더 = 첫 컬럼이 ▣ 포함하고 다른 칸 비어 있음
            if "▣" in col_cat and not col_url and not col_name:
                # 앞 공백·기호 정리
                current_section = col_cat.replace("▣", "").strip()
                # 일부 행은 "▣ ✅ ..." 처럼 ✅ 포함 — 그대로 보존
                continue

            if not col_url or not col_name:
                # URL·이름 둘 중 하나라도 없으면 스킵 (안내 행)
                continue

            order += 1
            rows.append({
                "sheet": sheet_name,
                "section": current_section or None,
                "category": col_cat or None,
                "name": col_name,
                "url": col_url,
                "description": col_desc or None,
                "search_text": make_search_text(col_name, col_desc, col_cat, current_section),
                "display_order": order,
            })

    return rows


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"ERROR: xlsx 파일 없음: {XLSX_PATH}", file=sys.stderr)
        return 1

    url, key = load_env()
    print(f"Supabase URL: {url.split('//')[1].split('.')[0]}...")

    rows = parse_xlsx()
    total = len(rows)
    print(f"Parsed {total} sites from xlsx")

    # 시트별 카운트
    sheet_counts: dict[str, int] = {}
    for r in rows:
        sheet_counts[r["sheet"]] = sheet_counts.get(r["sheet"], 0) + 1
    for s, n in sheet_counts.items():
        print(f"  {s}: {n}")

    if total == 0:
        print("⚠️ 파싱 결과 0건 — xlsx 구조 확인 필요", file=sys.stderr)
        return 2

    BATCH = 50
    # UNIQUE (sheet, name, url) → on_conflict 컬럼 3개 명시
    endpoint = "/rest/v1/monitoring_sites?on_conflict=sheet,name,url"
    prefer = "resolution=merge-duplicates,return=minimal"

    ok_count = 0
    for i in range(0, total, BATCH):
        chunk = rows[i:i + BATCH]
        status, body, _ = post(url, key, endpoint, chunk, prefer=prefer)
        if 200 <= status < 300:
            ok_count += len(chunk)
            print(f"  Batch {i // BATCH + 1:2d} ({ok_count}/{total}) OK")
        else:
            print(f"  Batch {i // BATCH + 1:2d} FAIL status={status}")
            print(f"    body: {str(body)[:500]}")
            return 3

    # 검증
    status, _, headers = post(
        url, key,
        "/rest/v1/monitoring_sites?select=id",
        None,
        prefer="count=exact,head=true",
    )
    cr = headers.get("Content-Range", "")
    if "/" in cr:
        actual = cr.split("/")[-1]
        print(f"\nFinal count (Content-Range): {actual}")
        if str(actual) == str(total):
            print("✅ 성공: 모든 row 적용됨")
            return 0
        else:
            print(f"⚠️ 불일치: 기대 {total}, 실제 {actual} (기존 row + 중복 정리 결과일 수 있음)")
            return 0
    print(f"\nVerification headers Content-Range 누락")
    return 0


if __name__ == "__main__":
    sys.exit(main())
